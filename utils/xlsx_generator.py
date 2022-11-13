from faker import Faker
import openpyxl
import os
import random

from utils.utils import get_flag
from utils import png_generator, jpg_generator
from utils.flags import flags

fake = Faker()

def generate_xlsx_file(path):
    if not os.path.exists(path):
        os.makedirs(path)
    if not os.path.exists("./temp/"):
        os.makedirs("./temp/")
    filename = fake.name().replace(" ", "_") + f"_{random.randint(1000,9999)}" + ".xlsx"
    flag = get_flag()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    insert_path = ""
    if random.randint(0, 1):
        ws['A1'] = flag
    else:
        insert_path, flag = random.choice([
            png_generator.generate_png_file,
            jpg_generator.generate_jpg_file
        ])("./temp/")
        img = openpyxl.drawing.image.Image(insert_path)
        # img.anchor(ws.cell(1,1))
        ws.add_image(img, "A1")
    wb.save(os.path.join(path, filename))
    if insert_path:
        os.remove(insert_path)
    return os.path.join(path, filename), flag

def generate_xlsx_files(path, count):
    for _ in range(count):
        flags.add(generate_xlsx_file(path))